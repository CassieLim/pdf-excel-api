from fastapi import FastAPI, UploadFile, File
from fastapi.responses import FileResponse
from fastapi.middleware.cors import CORSMiddleware

from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

import fitz  # PyMuPDF
from PIL import Image

import os
import io
import re
import shutil
from uuid import uuid4


app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["http://localhost:3000"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

BASE_DIR = os.path.dirname(__file__)
UPLOAD_DIR = os.path.join(BASE_DIR, "uploads")
OUTPUT_DIR = os.path.join(BASE_DIR, "outputs")

os.makedirs(UPLOAD_DIR, exist_ok=True)
os.makedirs(OUTPUT_DIR, exist_ok=True)


def block_lines(block):
    """把 PyMuPDF text block 拆成一行一行"""
    lines = []
    for line in block.get("lines", []):
        text = "".join(span["text"] for span in line.get("spans", [])).strip()
        if text:
            lines.append(text)
    return lines


def clean_product_name(name: str) -> str:
    """清理商品名称里的杂讯"""
    name = name.replace("↑", "").replace("￾", "")
    name = re.sub(r"\n{3,}", "\n\n", name)
    return name.strip()


def parse_pdf_rows(pdf_path: str):
    """
    解析 PDF：
    - 以左边图片 block 当作每一列商品的锚点
    - 抓同一列的文字
    - 最后 10 行当作数据栏位
    - 前面的行当作 Product name
    """
    doc = fitz.open(pdf_path)

    report_title = ""
    report_datetime = ""

    rows = []

    for page_index in range(len(doc)):
        page = doc[page_index]
        page_no = page_index + 1

        data = page.get_text("dict")
        blocks = data["blocks"]

        text_blocks = [b for b in blocks if b["type"] == 0]
        image_blocks = sorted(
            [b for b in blocks if b["type"] == 1 and b["bbox"][0] < 100],
            key=lambda b: b["bbox"][1]
        )

        if page_no == 1:
            for tb in text_blocks:
                text = "\n".join(block_lines(tb))
                if "Product Closing Summary" in text and not report_title:
                    report_title = text.strip()
                if "Report Generated DateTime" in text and not report_datetime:
                    m = re.search(r"Report Generated DateTime\s*(.*)", text)
                    if m:
                        report_datetime = m.group(1).strip()
                    else:
                        report_datetime = text.strip()

        # 用图片区块来定义每一列商品的垂直范围
        for i, ib in enumerate(image_blocks):
            x0, y0, x1, y1 = ib["bbox"]

            # 跳过页首表头上方的非商品图片
            if y0 < 60:
                continue

            top = 60 if i == 0 else (image_blocks[i - 1]["bbox"][3] + y0) / 2
            bottom = page.rect.height if i == len(image_blocks) - 1 else (y1 + image_blocks[i + 1]["bbox"][1]) / 2

            row_text_blocks = []
            for tb in text_blocks:
                tx0, ty0, tx1, ty1 = tb["bbox"]
                cy = (ty0 + ty1) / 2

                # 只抓这一列高度范围内、且在图片右边的文字
                if top <= cy < bottom and tx0 > 95:
                    row_text_blocks.append(tb)

            row_text_blocks = sorted(row_text_blocks, key=lambda b: (b["bbox"][0], b["bbox"][1]))

            all_lines = []
            for tb in row_text_blocks:
                all_lines.extend(block_lines(tb))

            all_lines = [x for x in all_lines if x.strip()]

            # 正常一列：最后 10 行是固定数据
            # 其余行是 Product name
            if len(all_lines) < 11:
                continue

            value_lines = all_lines[-10:]
            name_lines = all_lines[:-10]

            product_name = clean_product_name("\n".join(name_lines))

            style = value_lines[0]
            sub_style = value_lines[1]
            product_no = value_lines[2]
            unit_price = value_lines[3]
            total_quantity = value_lines[4]
            qty_paid = value_lines[5]
            qty_unconfirmed = value_lines[6]
            qty_canceled = value_lines[7]
            qty_balance = value_lines[8]
            paid_quantity = value_lines[9]

            rows.append({
                "page_no": page_no,
                "product_name": product_name,
                "style": style,
                "sub_style": sub_style,
                "product_no": product_no,
                "storage_spaces": "",  # 这份 PDF 实际上大多没有独立值
                "unit_price": unit_price,
                "total_quantity": total_quantity,
                "qty_paid": qty_paid,
                "qty_unconfirmed": qty_unconfirmed,
                "qty_canceled": qty_canceled,
                "qty_balance": qty_balance,
                "paid_quantity": paid_quantity,
                "image_bytes": ib["image"],
            })

    doc.close()
    return report_title, report_datetime, rows


def safe_number_or_text(value):
    """数字就转数字，不是数字就保留原文字"""
    if value is None:
        return None
    text = str(value).strip()
    if text == "":
        return None
    if text == "--":
        return "--"
    if re.fullmatch(r"-?\d+", text):
        return int(text)
    if re.fullmatch(r"-?\d+\.\d+", text):
        return float(text)
    return text


def build_excel(pdf_filename: str, report_title: str, report_datetime: str, rows: list, output_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Closing Summary"

    # 顶部资讯
    ws["A1"] = report_title or "Product Closing Summary"
    ws["A2"] = f"Report Generated DateTime: {report_datetime}" if report_datetime else "Report Generated DateTime:"
    ws["A3"] = f"Source file: {pdf_filename}"

    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"].font = Font(bold=False, size=11)
    ws["A3"].font = Font(italic=True, size=10)

    headers = [
        "Photo",
        "Product name",
        "Style",
        "Sub-style",
        "Product No.",
        "Storage spaces",
        "Unit price",
        "Total Quantity",
        "QTY (Paid)",
        "QTY (unconfirmed)",
        "QTY (canceled)",
        "QTY (Balance)",
        "Paid Quantity",
        "Source Page",
    ]

    header_row = 4
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # 欄宽
    widths = {
        "A": 14,
        "B": 34,
        "C": 12,
        "D": 12,
        "E": 16,
        "F": 16,
        "G": 12,
        "H": 14,
        "I": 12,
        "J": 18,
        "K": 15,
        "L": 14,
        "M": 14,
        "N": 12,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    first_photo_temp_path = None

    for idx, row in enumerate(rows, start=5):
        ws.row_dimensions[idx].height = 70

        # 先写文字
        ws.cell(idx, 2, row["product_name"])
        ws.cell(idx, 3, safe_number_or_text(row["style"]))
        ws.cell(idx, 4, safe_number_or_text(row["sub_style"]))
        ws.cell(idx, 5, safe_number_or_text(row["product_no"]))
        ws.cell(idx, 6, safe_number_or_text(row["storage_spaces"]))
        ws.cell(idx, 7, safe_number_or_text(row["unit_price"]))
        ws.cell(idx, 8, safe_number_or_text(row["total_quantity"]))
        ws.cell(idx, 9, safe_number_or_text(row["qty_paid"]))
        ws.cell(idx, 10, safe_number_or_text(row["qty_unconfirmed"]))
        ws.cell(idx, 11, safe_number_or_text(row["qty_canceled"]))
        ws.cell(idx, 12, safe_number_or_text(row["qty_balance"]))
        ws.cell(idx, 13, safe_number_or_text(row["paid_quantity"]))
        ws.cell(idx, 14, row["page_no"])

        # 对齐
        for col in range(2, 15):
            ws.cell(idx, col).alignment = Alignment(vertical="top", wrap_text=True)

        # 插入图片到 A 栏
        img_temp_path = os.path.join(OUTPUT_DIR, f"temp_{uuid4().hex}.png")
        Image.open(io.BytesIO(row["image_bytes"])).save(img_temp_path)

        xl_img = XLImage(img_temp_path)
        xl_img.width = 60
        xl_img.height = 60
        ws.add_image(xl_img, f"A{idx}")

        if idx == 5:
            first_photo_temp_path = img_temp_path

    # 冻结窗格
    ws.freeze_panes = "A5"

    # 第二个工作表：First Photo
    ws2 = wb.create_sheet("First Photo")
    ws2["A1"] = "First product photo"
    ws2["A2"] = f"Source file: {pdf_filename}"
    ws2["A3"] = "Product name"
    ws2["B3"] = rows[0]["product_name"] if rows else ""
    ws2["A4"] = "Product No."
    ws2["B4"] = rows[0]["product_no"] if rows else ""
    ws2["A5"] = "Source Page"
    ws2["B5"] = rows[0]["page_no"] if rows else ""

    ws2["A1"].font = Font(bold=True, size=14)
    ws2.column_dimensions["A"].width = 18
    ws2.column_dimensions["B"].width = 40
    ws2.row_dimensions[7].height = 180

    if first_photo_temp_path and os.path.exists(first_photo_temp_path):
        first_img = XLImage(first_photo_temp_path)
        first_img.width = 180
        first_img.height = 180
        ws2.add_image(first_img, "A7")

    wb.save(output_path)


@app.get("/")
def root():
    return {"message": "API is running"}


@app.post("/convert")
async def convert_pdf(file: UploadFile = File(...)):
    job_id = str(uuid4())
    safe_filename = file.filename.replace("/", "_").replace("\\", "_")
    pdf_path = os.path.join(UPLOAD_DIR, f"{job_id}_{safe_filename}")

    with open(pdf_path, "wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    report_title, report_datetime, rows = parse_pdf_rows(pdf_path)

    output_filename = f"{os.path.splitext(safe_filename)[0]}.xlsx"
    output_path = os.path.join(OUTPUT_DIR, f"{job_id}_{output_filename}")

    build_excel(
        pdf_filename=safe_filename,
        report_title=report_title,
        report_datetime=report_datetime,
        rows=rows,
        output_path=output_path,
    )

    return FileResponse(
        output_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=output_filename,
    )
    